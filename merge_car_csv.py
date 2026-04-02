import csv
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- 경로 설정 (사용자 환경에 맞게 수정) ---
# 기본 베이스 경로를 설정합니다.
BASE_DIR = r"C:\Users\정지원\OneDrive - GK리서치\바탕 화면\Crawler\RAG-Keyword-Matcher"

# 카테고리 이름 치환 규칙 (브랜드별로 다를 수 있음)
CATEGORY_RENAME = {
    "Driver Assistance Technology": "Drive Wise Technologies",
    "BATTERY ELECTRIC MOTOR": "Electric Motor Specs", # 예시
}

def normalize_category(cat):
    return CATEGORY_RENAME.get(cat, cat)

def load_url_map(brand):
    """configs/{brand}.json 파일에서 {model_key: url} 맵을 로드합니다."""
    config_path = os.path.join(BASE_DIR, "configs", f"{brand}.json")
    if not os.path.exists(config_path):
        print(f"경고: {config_path} 설정 파일이 없습니다. URL 없이 진행합니다.")
        return {}
    
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)
    return {k: v["url"] for k, v in config.get("models", {}).items()}

def model_display_name(model_raw):
    """'Carnival_Hybrid' -> 'Carnival Hybrid'"""
    return model_raw.replace("_", " ").title()

def sheet_name_from_filename(brand, filename):
    """'lexus_es.csv' -> 'Es' (브랜드 접두사 제거)"""
    name = filename.replace(f"{brand}_", "").replace(".csv", "")
    return name.replace("_", " ").title()

def read_csv(filepath):
    rows = []
    # lexus_es.csv 예시처럼 같은 태그가 섞여있을 경우를 대비해 처리 필요
    # 일반적인 csv.DictReader는 헤더에 불필요한 텍스트가 있으면 오동작할 수 있음
    with open(filepath, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        for row in reader:
            rows.append(row)
    return headers, rows

def get_trim_columns(headers):
    fixed = {"Brand", "Model", "Year", "Category", "Feature"}
    # 고정 헤더 이외의 모든 컬럼(트림명)을 반환
    return [h for h in headers if h not in fixed and h is not None]

# --- 스타일 함수 (기존 유지) ---
def apply_header_style(cell):
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_category_style(cell):
    cell.font = Font(bold=False, size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data_style(cell):
    cell.font = Font(size=9)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# --- 시트 생성 로직 (기존 유지) ---
def build_combined_sheet(ws, all_data):
    trim_keys = []
    seen_trims = set()
    for model, trim, cat, feat, val in all_data:
        key = (model, trim)
        if key not in seen_trims:
            trim_keys.append(key)
            seen_trims.add(key)

    feat_keys = []
    seen_feats = set()
    for model, trim, cat, feat, val in all_data:
        key = (cat, feat)
        if key not in seen_feats:
            feat_keys.append(key)
            seen_feats.add(key)
    feat_keys.sort(key=lambda x: x[0])

    lookup = {(m, t, c, f): v for m, t, c, f, v in all_data}

    ws.cell(row=1, column=1, value="카테고리")
    ws.cell(row=1, column=2, value="피처")
    apply_header_style(ws.cell(row=1, column=1))
    apply_header_style(ws.cell(row=1, column=2))

    for col_idx, (model, trim) in enumerate(trim_keys, start=3):
        header_val = f"{model} / {trim}"
        cell = ws.cell(row=1, column=col_idx, value=header_val)
        apply_header_style(cell)

    for row_idx, (cat, feat) in enumerate(feat_keys, start=2):
        apply_category_style(ws.cell(row=row_idx, column=1, value=cat))
        apply_data_style(ws.cell(row=row_idx, column=2, value=feat))
        for col_idx, (model, trim) in enumerate(trim_keys, start=3):
            val = lookup.get((model, trim, cat, feat), "")
            apply_data_style(ws.cell(row=row_idx, column=col_idx, value=val))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    for col_idx in range(3, len(trim_keys) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "C2"

def build_individual_sheet(ws, headers, rows, url=""):
    trim_cols = get_trim_columns(headers)
    year = rows[0].get("Year", "") if rows else ""
    
    ws.cell(row=1, column=1, value="Year")
    ws.cell(row=1, column=2, value=year)
    ws.cell(row=1, column=3, value="URL")
    ws.cell(row=1, column=4, value=url)
    
    HEADER_ROW = 3
    ws.cell(row=HEADER_ROW, column=1, value="카테고리")
    ws.cell(row=HEADER_ROW, column=2, value="피처")
    apply_header_style(ws.cell(row=HEADER_ROW, column=1))
    apply_header_style(ws.cell(row=HEADER_ROW, column=2))
    
    for col_idx, trim in enumerate(trim_cols, start=3):
        apply_header_style(ws.cell(row=HEADER_ROW, column=col_idx, value=trim))

    for row_idx, row in enumerate(rows, start=HEADER_ROW + 1):
        cat = normalize_category(row.get("Category", ""))
        feat = row.get("Feature", "")
        apply_category_style(ws.cell(row=row_idx, column=1, value=cat))
        apply_data_style(ws.cell(row=row_idx, column=2, value=feat))
        for col_idx, trim in enumerate(trim_cols, start=3):
            apply_data_style(ws.cell(row=row_idx, column=col_idx, value=row.get(trim, "")))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    for col_idx in range(3, len(trim_cols) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = f"C{HEADER_ROW + 1}"

def main(brand_name):
    # 경로 동적 설정
    csv_dir = os.path.join(BASE_DIR, "storage", "csv", brand_name)
    output_file = os.path.join(BASE_DIR, "storage", "csv", f"{brand_name}_combined.xlsx")
    
    if not os.path.exists(csv_dir):
        print(f"오류: '{csv_dir}' 경로를 찾을 수 없습니다.")
        return

    url_map = load_url_map(brand_name)
    csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith(".csv")])

    all_data = []
    file_data = []

    for filename in csv_files:
        filepath = os.path.join(csv_dir, filename)
        headers, rows = read_csv(filepath)
        trim_cols = get_trim_columns(headers)
        sheet_name = sheet_name_from_filename(brand_name, filename)
        
        model_key = filename.replace(f"{brand_name}_", "").replace(".csv", "")
        url = url_map.get(model_key, "")
        file_data.append((sheet_name, headers, rows, url))

        for row in rows:
            model = model_display_name(row.get("Model", ""))
            cat = normalize_category(row.get("Category", ""))
            feat = row.get("Feature", "")
            for trim in trim_cols:
                val = row.get(trim, "")
                all_data.append((model, trim, cat, feat, val))

    wb = Workbook()
    ws_combined = wb.active
    ws_combined.title = "전체 통합"
    build_combined_sheet(ws_combined, all_data)

    for sheet_name, headers, rows, url in file_data:
        ws = wb.create_sheet(title=sheet_name)
        build_individual_sheet(ws, headers, rows, url=url)

    wb.save(output_file)
    print(f"[{brand_name.upper()}] 저장 완료: {output_file}")

if __name__ == "__main__":
    # 실행 시 인자로 브랜드명을 넘기거나 (python merge_car_csv.py lexus)
    # 인자가 없으면 기본적으로 'lexus'를 수행하도록 설정
    target_brand = sys.argv[1] if len(sys.argv) > 1 else "lexus"
    main(target_brand)